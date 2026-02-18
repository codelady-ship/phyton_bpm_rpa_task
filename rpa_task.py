from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import time
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

EXCEL_FILE = "data.xlsx"
RESULT_DIR = "result"
RESULT_FILE = os.path.join(RESULT_DIR, "execution_log.xlsx")
URL = "https://practicetestautomation.com/practice-test-login/"

allowed_users_df = pd.read_excel(EXCEL_FILE)
allowed_users = dict(zip(
    allowed_users_df["username"].astype(str),
    allowed_users_df["password"].astype(str)
))

if not os.path.exists(RESULT_DIR):
    os.makedirs(RESULT_DIR)

driver = webdriver.Chrome()
driver.maximize_window()
wait = WebDriverWait(driver, 10)

execution_rows = []

test_users = [
    ("test1", "12345"),
    ("test2", "abcde"),
    ("someone", "wrongpass")
]

for username, password in test_users:
    try:
        driver.get(URL)

        username_input = wait.until(
            EC.presence_of_element_located((By.ID, "username"))
        )
        password_input = driver.find_element(By.ID, "password")

        username_input.clear()
        username_input.send_keys(username)
        password_input.clear()
        password_input.send_keys(password)

        driver.find_element(By.ID, "submit").click()
        time.sleep(2)

        if username in allowed_users and password == allowed_users[username]:
            result = "SUCCESS"
        else:
            result = "FAIL"

    except Exception as e:
        result = "SYSTEM_ERROR"
        print(f"System error for {username}: {e}")

    execution_rows.append({
        "username": username,
        "password": password,
        "result": result,
        "execution_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    })

try:
    if os.path.exists(RESULT_FILE):
        existing_df = pd.read_excel(RESULT_FILE)
        final_df = pd.concat([existing_df, pd.DataFrame(execution_rows)], ignore_index=True)
    else:
        final_df = pd.DataFrame(execution_rows)

    final_df.to_excel(RESULT_FILE, index=False)

except PermissionError:
    print("Excel faylı açıqdır. Zəhmət olmasa bağlayıb yenidən run et.")
    driver.quit()
    exit()

workbook = load_workbook(RESULT_FILE)
sheet = workbook.active
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

for row in range(2, sheet.max_row + 1):
    if sheet[f"C{row}"].value == "FAIL":
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=row, column=col).fill = red_fill

workbook.save(RESULT_FILE)

print("Automation finished. Results appended and FAIL rows highlighted.")

input("Press Enter to close the browser...")
driver.quit()
